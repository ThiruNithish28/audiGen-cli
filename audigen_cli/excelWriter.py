from openpyxl import load_workbook
from openpyxl.styles import Alignment
from audigen_cli import utils

def revisionHistoryChanges(wb, BRD_endDate):
  revision_sheet = wb['Revision History']
  revision_sheet['B8']=BRD_endDate
  revision_sheet['D8']=BRD_endDate

def addTestCases(wb, llm_generateTestCase):
  test_cases_sheet = wb['Test Cases']

  # -------------------------------
  # 1. Update Header Field (D4)
  # -------------------------------
  test_cases_sheet['D4'].value = (
      llm_generateTestCase.additional_info.componets_affected
  )

  # -------------------------------
  # 2. Write Test Cases
  # -------------------------------
  start_row = 8
  test_cases = llm_generateTestCase.testCases

  for i, test_case in enumerate(test_cases):
      row_num = start_row + i

      # Write values
      test_cases_sheet[f'A{row_num}'].value = test_case.testCaseId
      test_cases_sheet[f'B{row_num}'].value = test_case.testCase
      test_cases_sheet[f'C{row_num}'].value = test_case.expectedResult

      # -------------------------------
      # 3. Apply Formatting
      # -------------------------------
      for col in ['A', 'B', 'C']:
          cell = test_cases_sheet[f'{col}{row_num}']
          cell.alignment = Alignment(wrap_text=True, vertical='top')

      # Set row height (adjust as needed)
      # test_cases_sheet.row_dimensions[row_num].height = 60
      utils.auto_adjust_row_height(test_cases_sheet, row_num, ['A', 'B', 'C'])


  # -------------------------------
  # 4. Column Width (set once)
  # -------------------------------
  test_cases_sheet.column_dimensions['A'].width = 20
  test_cases_sheet.column_dimensions['B'].width = 45
  test_cases_sheet.column_dimensions['C'].width = 45

def updateImpactAnaylsis(wb,llm_generateTestCase,BRD_startDate,BRD_endDate):
  sheet = wb['Impact Analysis']
  effort = utils.calculateEffort(BRD_startDate,BRD_endDate)

  sheet['E4'].value = llm_generateTestCase.additional_info.brd_rasiedBy
  sheet['E7'].value = BRD_startDate
  sheet['B10'].value = llm_generateTestCase.additional_info.componets_affected
  sheet['F10'].value = effort
  sheet['G10'].value = BRD_startDate
  sheet['H10'].value = BRD_endDate
  sheet['I10'].value = effort
  sheet['E13'].value = BRD_endDate


def updateCodeCheckList(wb,llm_generateTestCase, BRD_endDate):
  sheet = wb['Java Checklist']
  sheet['B6'].value = llm_generateTestCase.additional_info.componets_affected
  sheet['D7'].value = BRD_endDate
  sheet['D8'].value = BRD_endDate

# -------------------------------
# main method
# -------------------------------
def startExcelChange(llm_generateTestCase,BRD_startDate,BRD_endDate):
    # Load template
    impact_analysis_xlsx = load_workbook('Impact Analysis.xlsx')
    test_case_xlsx = load_workbook('template/Test Cases.xlsx')
    code_checkList_xlsx = load_workbook('template/Code Review Checklist.xlsx')
    print(f"Images in template: {len(impact_analysis_xlsx['Impact Analysis']._images)}")
    # 1 impact analysis changes
    updateImpactAnaylsis(impact_analysis_xlsx, llm_generateTestCase,BRD_startDate,BRD_endDate)
    # 2 testcase xl
    revisionHistoryChanges(test_case_xlsx,BRD_endDate)
    addTestCases(test_case_xlsx,llm_generateTestCase)
    # 3 codeCheckList
    updateCodeCheckList(code_checkList_xlsx,llm_generateTestCase, BRD_endDate)

    #save
    test_case_xlsx.calculation.fullCalcOnLoad = True
    impact_analysis_xlsx.save('Impact Analysis Template.xlsx')
    test_case_xlsx.save('sampleCases.xlsx')
    code_checkList_xlsx.save('Code Checklist.xlsx')

    wb2 = load_workbook('Impact Analysis Template.xlsx')
    ws2 = wb2['Impact Analysis']
    print(f"Images after save: {len(ws2._images)}")
    print("All Excel files generated successfully")

